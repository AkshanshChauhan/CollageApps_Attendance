import openpyxl

my_wb = openpyxl.load_workbook("student.xlsx")
my_sheet = my_wb.active
rows = my_sheet.iter_rows(1,59)
cols = my_sheet.iter_cols(1, 9)

file = open("index.html", "w")

datastr = ""

for i in rows:
    for j in range(0,9):
        print("|",i[j].value, " |", end='')
        links = i[2].value
        datastr = datastr + str("%s <br>"%i[j].value)
    open("%s.html"%links, "w").write(datastr)
    file.write("<a href='%s.html'>Data of <b>%s</b></a><br>"%(links, links))
    datastr = ""
    print("")